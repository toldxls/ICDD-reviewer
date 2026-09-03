#!/usr/bin/env python3
"""
epma — microprobe data reduction: wt% oxides → empirical formula, the published EPMA table, and an
xlsx whose formulas show every step.

    python3 -m pxrd_review.epma <probe.xlsx|.csv|.txt> [--basis O=21 | cations=8 | U=5]
                                [--add H2O=structure:12 --add CO2=wt:14.02 --add H2O=difference]
                                [--convert UO2=UO3 --convert Tl2O3=Tl2O] [--drop Se] [--points 8-20]
                                [--standards CaO=anorthite,UO3=UO2+x] [--ideal "Ca3Na(UO2)(CO3)3(SeO3)F(H2O)6"]
                                [--journal ammin] [--word] [--xlsx] [--out DIR]
    pxrd epma probe.xlsx --basis O=21 --add CO2=structure:3 --add H2O=structure:6

Input: the probe's own export (JEOL/Cameca: a header row naming the oxides — 'CaO (2) Oxide%' or
plain 'CaO' — and one row per analysis point; a 'Line Numbers'/'Oxide Totals'/formula columns are
recognised and skipped), or a two-column constituent/wt% list, or a reduced table with mean, range
and s.d. columns.

Reduction (the way the owner's spreadsheets do it): for each constituent, moles = wt% / MW, cations
= moles × cations per formula, anions = moles × oxygens per formula (F, Cl count one anion each,
H2O counts one); the normalisation factor comes from the basis —
  O=N        N anions (O + F + Cl) per formula unit  (the usual 'on the basis of N O apfu')
  cations=N  N cations per formula unit (H excluded)
  El=N       N atoms of one element (U=5, Si+Al=4)
  charge     balance: the OH/H2O split, or Fe3+/Fe2+, adjusted until Σ(+) = Σ(−) on a fixed-O basis
Constituents that the probe cannot measure are added with --add: 'structure:N' (N apfu from the
ideal formula, converted to wt% at the current normalisation), 'wt:X' (a measured value, e.g. TGA
water), or 'difference' (100 − total). F=O / Cl=O corrections are applied; UO2→UO3-type
conversions with --convert. The published table lists constituent, mean, range, s.d., standard,
normalised and ideal wt%, F=O, total; the xlsx has the raw points, the reduction with live
formulas, and the table.
"""
import os, re, sys, math, argparse, statistics
from collections import OrderedDict, namedtuple

# IUPAC 2013 standard atomic weights (abridged), enough for mineral analyses
ATOMIC_WEIGHTS = {
    'H': 1.008, 'He': 4.0026, 'Li': 6.94, 'Be': 9.0122, 'B': 10.81, 'C': 12.011, 'N': 14.007, 'O': 15.999,
    'F': 18.998, 'Ne': 20.180, 'Na': 22.990, 'Mg': 24.305, 'Al': 26.982, 'Si': 28.085, 'P': 30.974,
    'S': 32.06, 'Cl': 35.45, 'Ar': 39.948, 'K': 39.098, 'Ca': 40.078, 'Sc': 44.956, 'Ti': 47.867,
    'V': 50.942, 'Cr': 51.996, 'Mn': 54.938, 'Fe': 55.845, 'Co': 58.933, 'Ni': 58.693, 'Cu': 63.546,
    'Zn': 65.38, 'Ga': 69.723, 'Ge': 72.630, 'As': 74.922, 'Se': 78.971, 'Br': 79.904, 'Kr': 83.798,
    'Rb': 85.468, 'Sr': 87.62, 'Y': 88.906, 'Zr': 91.224, 'Nb': 92.906, 'Mo': 95.95, 'Tc': 98.0,
    'Ru': 101.07, 'Rh': 102.91, 'Pd': 106.42, 'Ag': 107.87, 'Cd': 112.41, 'In': 114.82, 'Sn': 118.71,
    'Sb': 121.76, 'Te': 127.60, 'I': 126.90, 'Xe': 131.29, 'Cs': 132.91, 'Ba': 137.33, 'La': 138.91,
    'Ce': 140.12, 'Pr': 140.91, 'Nd': 144.24, 'Pm': 145.0, 'Sm': 150.36, 'Eu': 151.96, 'Gd': 157.25,
    'Tb': 158.93, 'Dy': 162.50, 'Ho': 164.93, 'Er': 167.26, 'Tm': 168.93, 'Yb': 173.05, 'Lu': 174.97,
    'Hf': 178.49, 'Ta': 180.95, 'W': 183.84, 'Re': 186.21, 'Os': 190.23, 'Ir': 192.22, 'Pt': 195.08,
    'Au': 196.97, 'Hg': 200.59, 'Tl': 204.38, 'Pb': 207.2, 'Bi': 208.98, 'Po': 209.0, 'Th': 232.04,
    'Pa': 231.04, 'U': 238.03, 'Np': 237.0, 'Pu': 244.0,
}
ANIONS = {'F': -1, 'Cl': -1, 'Br': -1, 'I': -1, 'S': -2, 'Se': -2, 'Te': -2, 'O': -2}

Constituent = namedtuple('Constituent', 'formula element n_cat n_o charge mw kind')
# kind: 'oxide' | 'element-anion' (F, Cl, S as sulfide) | 'water' | 'other'

_FORM = re.compile(r'([A-Z][a-z]?)(\d*)')

def parse_constituent(formula):
    """'Al2O3' -> Constituent(Al, 2 cations, 3 O, +3); 'F' -> anion; 'H2O' -> water; 'UO3' -> U6+.
    A qualifier in parentheses ('H2O(calc)', 'FeO(Mössbauer)') is ignored; 'Nb205' (a zero for
    the O) is read as Nb2O5."""
    f = re.sub(r'\([^)]*\)', '', formula.strip().replace('*', ''))
    f = re.sub(r'\s+', '', f)
    f = re.sub(r'^([A-Z][a-z]?\d?)0(\d)$', r'\1O\2', f)
    parts = _FORM.findall(f)
    if not parts or ''.join(a + b for a, b in parts) != f:
        raise ValueError('not a constituent formula: %r' % formula)
    counts = OrderedDict()
    for el, n in parts:
        if el not in ATOMIC_WEIGHTS:
            raise ValueError('unknown element %r in %r' % (el, formula))
        counts[el] = counts.get(el, 0) + (int(n) if n else 1)
    mw = sum(ATOMIC_WEIGHTS[el] * n for el, n in counts.items())
    if list(counts) == ['H', 'O'] and counts['H'] == 2 * counts['O']:
        return Constituent(f, 'H', counts['H'], counts['O'], 1, mw, 'water')
    if len(counts) == 1:
        el = next(iter(counts))
        if el in ANIONS:
            return Constituent(f, el, counts[el], 0, ANIONS[el], mw, 'element-anion')
        return Constituent(f, el, counts[el], 0, 0, mw, 'element')   # metal, valence unknown
    if 'O' in counts and len(counts) == 2:
        el = next(k for k in counts if k != 'O')
        n_cat, n_o = counts[el], counts['O']
        return Constituent(f, el, n_cat, n_o, 2 * n_o / n_cat, mw, 'oxide')
    # a hydroxide / carbonate-like constituent: charge from oxygens over the first element
    el = next(iter(counts)); n_o = counts.get('O', 0)
    return Constituent(f, el, counts[el], n_o, 2 * n_o / counts[el] if counts[el] else 0, mw, 'other')

# ----------------------------------------------------------------------------- reading the probe export

Dataset = namedtuple('Dataset', 'constituents points labels standards source reduced')
# reduced: None, or {'mean':[], 'sd':[], 'lo':[], 'hi':[], 'n': int} when the input is already a table

_OX_HEAD = re.compile(r'^\s*([A-Z][a-z]?\d*O\d*|H2O|F|Cl|Br|S|Se|Te|[A-Z][a-z]?)\b')

def _cells_from_file(path):
    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext in ('.xlsx', '.xlsm'):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            sheet = [list(r) for r in ws.iter_rows(values_only=True)]
            rows.append((ws.title, sheet))
    else:
        import csv
        with open(path, encoding='utf-8-sig', errors='replace') as f:
            text = f.read()
        dialect = 'excel-tab' if text.count('\t') > text.count(',') else 'excel'
        sheet = [[(_num_or_str(c)) for c in r] for r in csv.reader(text.splitlines(), dialect=dialect)]
        rows.append((os.path.basename(path), sheet))
    return rows

def _num_or_str(c):
    try:
        return float(c)
    except (TypeError, ValueError):
        return c

def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def load_probe(path, sheet=None):
    """Find the block: a header row with >= 2 constituent names, then numeric rows below it."""
    best = None
    for title, grid in _cells_from_file(path):
        if sheet and title != sheet:
            continue
        for ri, row in enumerate(grid):
            heads = {}
            for ci, cell in enumerate(row):
                if not isinstance(cell, str):
                    continue
                s = cell.strip()
                if re.search(r'formula|total|line|number|point|sample|comment', s, re.I):
                    continue
                s = s.replace('(NH4)2O', 'N2H8O')                          # ammonium as its 'oxide'
                m = re.match(r'^(N2H8O|[A-Z][a-z]?\d*O\d*|H2O|CO2|F|Cl|Br|S|Se|Te|[A-Z][a-z]?)(?![a-z])', s)
                if m and (m.group(1) not in ('O', 'H', 'C', 'N') or m.group(1) == 'N2H8O'):
                    try:
                        parse_constituent(m.group(1)); heads[ci] = m.group(1)
                    except ValueError:
                        pass
            if len(heads) < 1:
                continue
            # numeric rows below, until a blank / non-numeric run
            pts, labels = [], []; gap = 0
            first_col = min(heads)
            for r2 in grid[ri + 1:]:
                vals = [r2[ci] if ci < len(r2) else None for ci in heads]
                if sum(_is_num(v) for v in vals) >= max(1, len(heads) // 2):
                    pts.append([float(v) if _is_num(v) else float('nan') for v in vals]); gap = 0
                    lab = r2[first_col - 1] if first_col >= 1 and len(r2) > first_col - 1 and r2[first_col - 1] is not None else len(pts)
                    labels.append(str(lab).strip() if isinstance(lab, str) else ('%g' % lab if _is_num(lab) else str(lab)))
                elif pts:
                    gap += 1                                    # a comment / blank line inside the block
                    if gap > 6 or any(isinstance(v, str) and re.search(r'average|mean|std|s\.d', v, re.I) for v in vals):
                        break
            if len(pts) >= 1 and (best is None or len(pts) * len(heads) > len(best[1]) * len(best[0])):
                stds = _standards_from(grid)
                best = (list(heads.values()), pts, labels, stds, '%s:%s' % (os.path.basename(path), title))
    if best is None:
        raise ValueError('no probe data found in %s (a header row naming the oxides, points below it)' % path)
    heads, pts, labels, stds, src = best
    # an already-reduced table? (columns Mean / S.D. / Range beside constituents in rows)
    return Dataset([parse_constituent(h) for h in heads], pts, labels, stds, src, None)

def _standards_from(grid):
    """'Stds: anorthite, albite, UO2' lines the JEOL export writes -> {element: standard}"""
    out = {}
    for row in grid[:12]:
        for cell in row:
            if isinstance(cell, str) and re.search(r'\bstds?\s*:', cell, re.I):
                tail = re.split(r'\bstds?\s*:', cell, flags=re.I)[1]
                for tok in re.split(r'[,;]\s*', tail):
                    tok = tok.strip()
                    if tok:
                        out[len(out)] = tok
    return out

# ----------------------------------------------------------------------------- reduction

Row = namedtuple('Row', 'c wt mean sd lo hi n mol cations anions apfu o_apfu source')

def _stats(vals):
    v = [x for x in vals if x == x]
    if not v:
        return (float('nan'),) * 4 + (0,)
    return (sum(v) / len(v), statistics.stdev(v) if len(v) > 1 else 0.0, min(v), max(v), len(v))

def reduce(ds, basis=('O', 21), adds=(), converts=(), drop=(), points=None, ideal_apfu=None, raw_anions=False):
    """-> Reduction. basis: ('O', N) | ('cations', N) | ('element', 'U', N) | ('charge', base) where
    base is one of the others and the balance adjusts the H2O/OH (or Fe3+/Fe2+) content.
    adds: [(formula, mode, value)] with mode 'structure' (apfu), 'wt' (wt%), 'difference'.
    converts: [('UO2', 'UO3')]. drop: constituents to leave out."""
    sel = range(len(ds.points)) if points is None else points
    cons = list(ds.constituents); table = OrderedDict()
    for j, c in enumerate(cons):
        if c.formula in drop:
            continue
        vals = [ds.points[i][j] for i in sel]
        mean, sd, lo, hi, n = _stats(vals)
        table[c.formula] = {'c': c, 'mean': mean, 'sd': sd, 'lo': lo, 'hi': hi, 'n': n, 'source': 'measured'}
    # conversions (UO2 -> UO3: wt% scales by MW ratio per cation)
    for a, b in converts:
        if a in table:
            ca, cb = table[a]['c'], parse_constituent(b)
            k = (cb.mw / cb.n_cat) / (ca.mw / ca.n_cat)
            rec = table[a]
            rec = dict(rec, c=cb, mean=rec['mean'] * k, sd=rec['sd'] * k, lo=rec['lo'] * k, hi=rec['hi'] * k,
                       source='converted from %s' % a)
            table = OrderedDict((b if key == a else key, rec if key == a else val) for key, val in table.items())
    # unmeasured constituents
    diff_key = None
    for formula, mode, value in adds:
        c = parse_constituent(formula)
        if mode == 'wt':
            table[formula] = {'c': c, 'mean': float(value), 'sd': 0.0, 'lo': float(value), 'hi': float(value), 'n': 0, 'source': 'given'}
        elif mode == 'difference':
            table[formula] = {'c': c, 'mean': 0.0, 'sd': 0.0, 'lo': 0.0, 'hi': 0.0, 'n': 0, 'source': 'by difference'}
            diff_key = formula
        elif mode == 'structure':
            table[formula] = {'c': c, 'mean': 0.0, 'sd': 0.0, 'lo': 0.0, 'hi': 0.0, 'n': 0, 'source': 'by structure (%s apfu)' % value,
                              'struct_apfu': float(value)}
    def anion_correction():
        # F=O, Cl=O: the oxygen already counted in the oxides that the halogen replaces
        corr = 0.0
        for k, r in table.items():
            if r['c'].kind == 'element-anion' and r['c'].element in ('F', 'Cl', 'Br', 'I'):
                corr += r['mean'] * ATOMIC_WEIGHTS['O'] / ATOMIC_WEIGHTS[r['c'].element] * 0.5
            if r['c'].kind == 'element-anion' and r['c'].element == 'S' and any(x['c'].kind == 'oxide' for x in table.values()):
                corr += r['mean'] * ATOMIC_WEIGHTS['O'] / ATOMIC_WEIGHTS['S']
        return corr
    # iterate: structure-based additions depend on the normalisation, which depends on the total
    for _ in range(12):
        rows = _apfu(table, basis, raw_anions)
        changed = False
        for k, r in table.items():
            if 'struct_apfu' in r:
                # wt% that gives struct_apfu at the current factor: apfu = mol × n_cat × factor
                fac = rows['factor']
                if fac > 0:
                    target = r['struct_apfu'] / (r['c'].n_cat * fac) * r['c'].mw
                    if abs(target - r['mean']) > 1e-6:
                        r['mean'] = r['lo'] = r['hi'] = target; changed = True
        if diff_key:
            total_other = sum(r['mean'] for k, r in table.items() if k != diff_key) - anion_correction()
            target = max(0.0, 100.0 - total_other)
            if abs(target - table[diff_key]['mean']) > 1e-6:
                table[diff_key]['mean'] = table[diff_key]['lo'] = table[diff_key]['hi'] = target; changed = True
        if not changed:
            break
    rows = _apfu(table, basis, raw_anions)
    out = OrderedDict()
    for k, r in table.items():
        out[k] = Row(r['c'], r['mean'], r['mean'], r['sd'], r['lo'], r['hi'], r['n'], rows['mol'][k], rows['cat'][k], rows['an'][k], rows['apfu'][k], rows['o'][k], r['source'])
    corr = anion_correction()
    total = sum(r.wt for r in out.values()) - corr
    return Reduction(out, basis, rows['factor'], corr, total, rows['charge'], ds)

class Reduction:
    def __init__(self, rows, basis, factor, corr, total, charge, ds):
        self.rows, self.basis, self.factor, self.corr, self.total, self.charge, self.ds = rows, basis, factor, corr, total, charge, ds
        self.anions_structure = None
    def formula(self, order=None, digits=2):
        """A plain empirical formula: cations in the given (or table) order, then O and (OH)/H2O."""
        parts = []
        keys = order or [k for k in self.rows]
        h2o = 0.0; o_tot = 0.0; anions = []
        for k in keys:
            r = self.rows.get(k)
            if not r:
                continue
            c = r.c
            if c.kind == 'water':
                h2o += r.apfu / 2; continue
            if c.kind == 'element-anion':
                anions.append((c.element, r.apfu)); continue
            n = ('%.*f' % (digits, r.apfu)).rstrip('0').rstrip('.')
            parts.append('%s%s' % (c.element, '' if n == '1' else n))
        o = self.o_apfu()
        s = ''.join(parts)
        if o:
            s += 'O%s' % ('%.*f' % (digits, o)).rstrip('0').rstrip('.')
        for el, n in anions:
            s += '%s%s' % (el, ('%.*f' % (digits, n)).rstrip('0').rstrip('.'))
        if h2o:
            s += '·%sH2O' % ('%.*f' % (digits, h2o)).rstrip('0').rstrip('.')
        return s
    def o_apfu(self):
        """O apfu of the oxides (the water oxygen is written as H2O)."""
        return sum(r.o_apfu for r in self.rows.values() if r.c.kind in ('oxide', 'other'))
    def sum_cations(self):
        return sum(r.apfu for r in self.rows.values() if r.c.kind in ('oxide', 'other', 'element') and r.c.element != 'H')

def _apfu(table, basis, raw_anions=False):
    mol = {k: (r['mean'] / r['c'].mw if r['c'].mw else 0.0) for k, r in table.items()}
    cat = {k: mol[k] * r['c'].n_cat for k, r in table.items()}
    an = {}
    for k, r in table.items():
        c = r['c']
        an[k] = mol[k] * c.n_o if c.kind in ('oxide', 'other') else (mol[k] * c.n_cat if c.kind == 'element-anion' else (mol[k] if c.kind == 'water' else 0.0))
    kind, *arg = basis
    if kind == 'O':
        # every anion: O of the oxides and of H2O, plus F, Cl (S). A halogen REPLACES an oxygen that the
        # oxide wt% still carries, so the oxide O is reduced by F/2 (Cl/2, S/1) — the O=F correction in
        # moles — unless raw_anions, the spreadsheet convention that counts both and picks N accordingly
        tot = sum(an.values())
        if not raw_anions:
            tot -= sum(an[k] * (0.5 if table[k]['c'].element in ('F', 'Cl', 'Br', 'I') else 1.0)
                       for k in table if table[k]['c'].kind == 'element-anion' and any(x['c'].kind == 'oxide' for x in table.values()))
        factor = arg[0] / tot if tot else 0.0
    elif kind == 'cations':
        tot = sum(v for k, v in cat.items() if table[k]['c'].element != 'H' and table[k]['c'].kind != 'element-anion')
        factor = arg[0] / tot if tot else 0.0
    elif kind == 'element':
        els = [e.strip() for e in arg[0].split('+')]
        tot = sum(v for k, v in cat.items() if table[k]['c'].element in els)
        factor = arg[1] / tot if tot else 0.0
    else:
        raise ValueError('unknown basis %r' % (basis,))
    apfu = {k: cat[k] * factor for k in table}
    o = {k: an[k] * factor for k in table}
    if kind == 'O' and not raw_anions:
        # report the oxide oxygen net of the halogen replacement, so O + F sums to the basis
        hal = sum(o[k] * (0.5 if table[k]['c'].element in ('F', 'Cl', 'Br', 'I') else 1.0)
                  for k in table if table[k]['c'].kind == 'element-anion')
        ox_o = sum(o[k] for k in table if table[k]['c'].kind in ('oxide', 'other'))
        if ox_o and hal:
            for k in table:
                if table[k]['c'].kind in ('oxide', 'other'):
                    o[k] *= (ox_o - hal) / ox_o
    # Σ(+): cations (H of water counts +1 each); Σ(−): 2 per O (oxides and water), 1 per F/Cl, 2 per S
    charge = sum(apfu[k] * table[k]['c'].charge for k in table if table[k]['c'].kind in ('oxide', 'other', 'water')) \
        + sum(apfu[k] * table[k]['c'].charge for k in table if table[k]['c'].kind == 'element-anion') \
        - 2 * sum(o[k] for k in table if table[k]['c'].kind in ('oxide', 'other', 'water'))
    return {'mol': mol, 'cat': cat, 'an': an, 'apfu': apfu, 'o': o, 'factor': factor, 'charge': charge}

def charge_balance(ds, base_basis, adds=(), converts=(), drop=(), points=None, adjust='H2O', anions=None, raw_anions=False):
    """Charge balance.
    adjust='H2O' — the hydrogen the analysis lacks: cations are normalised on a CATION or element
      basis (water is charge-neutral, so no anion basis can fix it), the structure's total anions per
      formula unit is `anions`, and H = 2·O_total + F − Σ(cation charges); it is reported as H2O.
    adjust='Fe' — total Fe (as FeO) is split into FeO + Fe2O3 until the charge balances on the given
      basis (bisection on the Fe3+ fraction)."""
    if adjust == 'H2O':
        if base_basis[0] == 'O':
            raise ValueError('charge balance for H2O needs a cation or element basis (cations=N or El=N) plus --anions N: '
                             'water is charge-neutral, so an anion basis cannot fix it')
        if not anions:
            raise ValueError('charge balance for H2O needs --anions N (total O + F per formula unit from the structure)')
        others = [t for t in adds if t[0] != 'H2O']
        r = reduce(ds, base_basis, others, converts, drop, points, raw_anions=raw_anions)
        hal = sum(x.apfu for x in r.rows.values() if x.c.kind == 'element-anion' and x.c.element in ('F', 'Cl', 'Br', 'I'))
        plus = sum(x.apfu * x.c.charge for x in r.rows.values() if x.c.kind in ('oxide', 'other', 'element'))
        h = 2 * (anions - hal) + hal - plus                       # Σ(−) − Σ(+), the H apfu needed
        h = max(h, 0.0)
        wt = h / 2 * (2 * ATOMIC_WEIGHTS['H'] + ATOMIC_WEIGHTS['O']) / r.factor if r.factor else 0.0
        out = reduce(ds, base_basis, others + [('H2O', 'wt', wt)], converts, drop, points, raw_anions=raw_anions)
        out.anions_structure = anions                       # the balance was struck against this total
        o_rep = out.o_apfu() + (out.rows['H2O'].o_apfu if 'H2O' in out.rows else 0.0)
        out.charge = out.charge + 2 * o_rep - 2 * (anions - hal)   # Σ(+) − Σ(−) against the structural anions
        return out
    if adjust == 'Fe':
        # Fe2+/Fe3+: on a cation (or element) basis with the structure's anion total, the charge the
        # analysis lacks with all Fe as Fe2+ is made up one unit per Fe3+ — a direct solve
        if base_basis[0] == 'O':
            raise ValueError('the Fe2+/Fe3+ split needs a cation or element basis (cations=N or El=N) plus --anions N')
        if not anions:
            raise ValueError('the Fe2+/Fe3+ split needs --anions N (total O + F per formula unit from the structure)')
        fe_keys = [c.formula for c in ds.constituents if c.element == 'Fe']
        if not fe_keys:
            raise ValueError('no Fe constituent to split')
        base = reduce(ds, base_basis, adds, converts, drop, points, raw_anions=raw_anions)
        feo_tot = sum(base.rows[k].wt * (71.844 / (base.rows[k].c.mw / base.rows[k].c.n_cat)) for k in fe_keys)   # all Fe as FeO wt%
        def run(x):
            extra = [('FeO', 'wt', feo_tot * (1 - x)), ('Fe2O3', 'wt', feo_tot * x * (159.687 / (2 * 71.844)))]
            r = reduce(ds, base_basis, list(adds) + extra, converts, list(drop) + fe_keys, points, raw_anions=raw_anions)
            r.anions_structure = anions
            hal = sum(v.apfu for v in r.rows.values() if v.c.kind == 'element-anion' and v.c.element in ('F', 'Cl', 'Br', 'I'))
            plus = sum(v.apfu * v.c.charge for v in r.rows.values() if v.c.kind in ('oxide', 'other', 'element', 'water'))
            r.charge = plus - (2 * (anions - hal) + hal)
            return r
        r0 = run(0.0)
        fe = sum(r0.rows[k].apfu for k in ('FeO', 'Fe2O3') if k in r0.rows)
        x = min(1.0, max(0.0, -r0.charge / fe)) if fe else 0.0   # each Fe3+ adds +1
        return run(x)
    raise ValueError('adjust must be H2O or Fe')

# ----------------------------------------------------------------------------- outputs

def _fmt(v, d=2):
    return '' if v is None or v != v else ('%.*f' % (d, v))

def published_table(red, standards=None, ideal_wt=None, name=''):
    """Rows for 'Constituent | Mean | Range | S.D. | Standard | Normalized | Ideal'."""
    standards = standards or {}
    norm_total = red.total
    head = ['Constituent', 'Mean', 'Range', 'S.D.', 'Standard', 'Normalized'] + (['Ideal'] if ideal_wt else [])
    rows = []
    for k, r in red.rows.items():
        c = r.c
        rng = '' if r.n < 2 else ('%s–%s' % (_fmt(r.lo), _fmt(r.hi)))
        sd = '' if r.n < 2 else _fmt(r.sd)
        std = standards.get(k, standards.get(c.element, ''))
        norm = _fmt(r.wt / norm_total * 100) if norm_total else ''
        row = [k + ('*' if r.source.startswith('by structure') or r.source == 'by difference' else ''), _fmt(r.wt), rng, sd, std, norm]
        if ideal_wt:
            row.append(_fmt(ideal_wt.get(k, 0.0)))
        rows.append(row)
    if red.corr:
        rows.append(['O=F,Cl' if any(r.c.element in ('F', 'Cl') for r in red.rows.values()) else 'O=S', _fmt(-red.corr), '', '', '', _fmt(-red.corr / norm_total * 100) if norm_total else ''] + ([''] if ideal_wt else []))
    rows.append(['Total', _fmt(red.total), '', '', '', '100.00'] + ([_fmt(sum(ideal_wt.values()))] if ideal_wt else []))
    notes = []
    for k, r in red.rows.items():
        if r.source.startswith('by structure') or r.source == 'by difference':
            notes.append('%s calculated %s.' % (k, r.source.replace('by structure', 'from the structure').replace('by difference', 'by difference')))
    return {'head': head, 'rows': rows, 'note': ' '.join(notes),
            'caption': 'Chemical composition (wt%%) of %s.' % name if name else 'Chemical composition (wt%).'}

def ideal_wt_percent(ideal_apfu):
    """{constituent: apfu} of the ideal formula -> {constituent: wt%} (each constituent's mass per
    formula unit, normalised to 100)."""
    mass = OrderedDict()
    for k, n in ideal_apfu.items():
        c = parse_constituent(k)
        mass[k] = n / c.n_cat * c.mw if c.kind != 'water' else n * c.mw
    tot = sum(mass.values())
    return OrderedDict((k, m / tot * 100) for k, m in mass.items())

def report_text(red, table):
    L = ['EPMA reduction — %s' % red.ds.source,
         '  basis: %s   normalisation factor %.5f   charge balance %+.3f%s'
         % (_basis_label(red.basis), red.factor, red.charge,
            ' (against %g anions per formula unit from the structure)' % red.anions_structure if red.anions_structure else ''), '']
    L.append('  %-10s %8s %8s %6s %10s %9s %9s %9s %9s   %s' % ('constituent', 'wt%', 's.d.', 'n', 'MW', 'moles', 'cations', 'apfu', 'O apfu', 'source'))
    for k, r in red.rows.items():
        L.append('  %-10s %8.2f %8.2f %6d %10.3f %9.5f %9.5f %9.4f %9.4f   %s' % (k, r.wt, r.sd, r.n, r.c.mw, r.mol, r.cations, r.apfu, r.o_apfu, r.source))
    if red.corr:
        L.append('  %-10s %8.2f' % ('O=F,Cl', -red.corr))
    L.append('  %-10s %8.2f' % ('total', red.total))
    L.append('')
    L.append('  empirical formula (%s): %s' % (_basis_label(red.basis), red.formula()))
    L.append('')
    L.append('Table. ' + table['caption'])
    w = [max(len(str(r[i])) for r in [table['head']] + table['rows']) for i in range(len(table['head']))]
    for r in [table['head']] + table['rows']:
        L.append('  ' + '  '.join(str(x).ljust(w[i]) for i, x in enumerate(r)).rstrip())
    if table['note']:
        L.append('  ' + table['note'])
    return '\n'.join(L)

def _basis_label(b):
    return {'O': '%s anions apfu', 'cations': '%s cations apfu'}.get(b[0], '%s').replace('%s', str(b[1]) if len(b) > 1 else '') if b[0] != 'element' else '%s = %s apfu' % (b[1], b[2])

def write_xlsx(red, table, path, ds=None, method=''):
    """raw points | reduction with live formulas | the published table | method (the paper's own
    statements of basis and treatment, when the inputs came from a paper)."""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter as L
    wb = openpyxl.Workbook()
    ds = ds or red.ds
    # ---- raw
    ws = wb.active; ws.title = 'raw'
    ws.append(['point'] + [c.formula for c in ds.constituents] + ['total'])
    for lab, pt in zip(ds.labels, ds.points):
        ws.append([lab] + [None if v != v else v for v in pt] + [sum(v for v in pt if v == v)])
    n = len(ds.points); ncol = len(ds.constituents)
    r_mean = n + 3
    ws.cell(r_mean, 1, 'mean'); ws.cell(r_mean + 1, 1, 's.d.'); ws.cell(r_mean + 2, 1, 'min'); ws.cell(r_mean + 3, 1, 'max')
    for j in range(ncol):
        col = L(j + 2); rng = '%s2:%s%d' % (col, col, n + 1)
        ws.cell(r_mean, j + 2, '=AVERAGE(%s)' % rng); ws.cell(r_mean + 1, j + 2, '=STDEV(%s)' % rng)
        ws.cell(r_mean + 2, j + 2, '=MIN(%s)' % rng); ws.cell(r_mean + 3, j + 2, '=MAX(%s)' % rng)
    # ---- reduction
    wr = wb.create_sheet('reduction')
    wr.append(['constituent', 'wt%', 'MW', 'moles', 'cations/unit', 'O/unit', 'cations', 'anions', 'apfu', 'O apfu', 'charge/cation', 'source'])
    for c in range(1, 13):
        wr.cell(1, c).font = Font(bold=True)
    raw_col = {c.formula: L(j + 2) for j, c in enumerate(ds.constituents)}
    first = 2
    for i, (k, r) in enumerate(red.rows.items()):
        row = first + i
        wr.cell(row, 1, k)
        if r.source == 'measured' and k in raw_col:
            wr.cell(row, 2, "=raw!%s%d" % (raw_col[k], r_mean))
        else:
            wr.cell(row, 2, r.wt)
        wr.cell(row, 3, r.c.mw); wr.cell(row, 4, '=B%d/C%d' % (row, row))
        n_an = r.c.n_o if r.c.kind in ('oxide', 'other') else (r.c.n_cat if r.c.kind == 'element-anion' else (1 if r.c.kind == 'water' else 0))
        wr.cell(row, 5, r.c.n_cat); wr.cell(row, 6, n_an)
        wr.cell(row, 7, '=D%d*E%d' % (row, row)); wr.cell(row, 8, '=D%d*F%d' % (row, row))
        wr.cell(row, 9, '=G%d*$B$%d' % (row, first + len(red.rows) + 4)); wr.cell(row, 10, '=H%d*$B$%d' % (row, first + len(red.rows) + 4) if r.c.kind in ('oxide', 'other', 'water') else 0)
        wr.cell(row, 11, r.c.charge); wr.cell(row, 12, r.source)
    last = first + len(red.rows) - 1
    rr = last + 1
    wr.cell(rr, 1, 'O=F,Cl'); wr.cell(rr, 2, -red.corr)
    wr.cell(rr + 1, 1, 'total'); wr.cell(rr + 1, 2, '=SUM(B%d:B%d)+B%d' % (first, last, rr))
    wr.cell(rr + 2, 1, 'Σ anions (for the basis)')
    kind = red.basis[0]
    if kind == 'O':
        wr.cell(rr + 2, 2, '=SUM(H%d:H%d)' % (first, last))
        wr.cell(rr + 3, 1, 'basis (anions apfu)'); wr.cell(rr + 3, 2, red.basis[1])
        wr.cell(rr + 4, 1, 'normalisation factor'); wr.cell(rr + 4, 2, '=B%d/B%d' % (rr + 3, rr + 2))
    elif kind == 'cations':
        wr.cell(rr + 2, 2, '=SUM(G%d:G%d)-SUMIF(A%d:A%d,"H2O",G%d:G%d)' % (first, last, first, last, first, last))
        wr.cell(rr + 3, 1, 'basis (cations apfu)'); wr.cell(rr + 3, 2, red.basis[1])
        wr.cell(rr + 4, 1, 'normalisation factor'); wr.cell(rr + 4, 2, '=B%d/B%d' % (rr + 3, rr + 2))
    else:
        wr.cell(rr + 2, 2, sum(r.cations for k, r in red.rows.items() if r.c.element in red.basis[1].split('+')))
        wr.cell(rr + 3, 1, 'basis (%s apfu)' % red.basis[1]); wr.cell(rr + 3, 2, red.basis[2])
        wr.cell(rr + 4, 1, 'normalisation factor'); wr.cell(rr + 4, 2, '=B%d/B%d' % (rr + 3, rr + 2))
    assert rr + 4 == first + len(red.rows) + 4
    wr.cell(rr + 5, 1, 'charge balance Σ(+) − Σ(−)'); wr.cell(rr + 5, 2, '=SUMPRODUCT(I%d:I%d,K%d:K%d)-2*SUM(J%d:J%d)' % (first, last, first, last, first, last))
    wr.cell(rr + 6, 1, 'empirical formula'); wr.cell(rr + 6, 2, red.formula())
    for col, w in zip('ABCDEFGHIJKL', (14, 10, 10, 10, 12, 8, 10, 10, 10, 10, 12, 26)):
        wr.column_dimensions[col].width = w
    # ---- table
    wt = wb.create_sheet('table')
    wt.append(table['head'])
    for c in range(1, len(table['head']) + 1):
        wt.cell(1, c).font = Font(bold=True)
    for r in table['rows']:
        wt.append(r)
    if table['note']:
        wt.append([]); wt.append([table['note']])
    if method:
        wm = wb.create_sheet('method')
        wm.append(['What the paper states (basis, calculated constituents) — the reduction sheet applies it']); wm.cell(1, 1).font = Font(bold=True)
        for part in method.split(' | '):
            wm.append([part])
        wm.append([]); wm.append(['basis applied', _basis_label(red.basis)])
        wm.column_dimensions['A'].width = 120
    wb.save(path)
    return path

# ----------------------------------------------------------------------------- ICDD entries: replicate the published formula
# An ICDD entry's Analysis field carries the mean wt% list and the empirical formula in the ICDD
# notation: "Microprobe analysis, average of 8 (wt.%): K2O 4.41, MnO 12.49, … H2O(calc) 27.13:
# ( Mn1.75 +2 Mg0.25 )sigma2.00 ( Fe1.84 +3 Al0.13 )sigma3.00 ( P O4 )4 [ O0.94 F0.81 ( O H )0.25 ]sigma2.00 !3.9 H2 O".
# Re-reducing the means on the paper's basis reproduces the coefficients — or shows which do not
# follow from the numbers (a constituent missing from the list, a value dropped, a typo, a slip
# in the formula). Group sums (Σ2.00) are checked against their parts as well.
_ICDD_WT = re.compile(r'(?<![A-Za-z])(N2H8O|H2O|CO2|SO3|[A-Z][a-z]?\d*O\d*|[A-Z][a-z]?)(?:\s*\([^)]{1,25}\))?[,:]?\s*(\d+\.\d+|\d+)(?![\d.]*[-–])\b')
_ICDD_TOK = re.compile(r'S\?|\(|\)|\[|\]|\{|\}|!|Σ|sigma|Sigma|SIGMA|[A-Z][a-z]?|\d+\.\d+|\d+|[+-]\d+(?:\.\d+)?|\?|·|\.|,|\s+')
CALCULATED_ELEMENTS = {'H', 'Li', 'B', 'Be', 'C', 'N'}          # what a probe cannot give: the authors calculate them
_USUAL_OXIDE = {'H': 'H2O', 'Li': 'Li2O', 'Na': 'Na2O', 'K': 'K2O', 'Rb': 'Rb2O', 'Cs': 'Cs2O', 'Tl': 'Tl2O', 'Ag': 'Ag2O', 'Cu': 'CuO',
                'Be': 'BeO', 'Mg': 'MgO', 'Ca': 'CaO', 'Sr': 'SrO', 'Ba': 'BaO', 'Mn': 'MnO', 'Fe': 'FeO', 'Co': 'CoO', 'Ni': 'NiO',
                'Zn': 'ZnO', 'Cd': 'CdO', 'Hg': 'HgO', 'Pb': 'PbO', 'B': 'B2O3', 'Al': 'Al2O3', 'Cr': 'Cr2O3', 'Ga': 'Ga2O3',
                'Bi': 'Bi2O3', 'Y': 'Y2O3', 'Sc': 'Sc2O3', 'La': 'La2O3', 'Ce': 'Ce2O3', 'Pr': 'Pr2O3', 'Nd': 'Nd2O3', 'Sm': 'Sm2O3',
                'Eu': 'Eu2O3', 'Gd': 'Gd2O3', 'Tb': 'Tb2O3', 'Dy': 'Dy2O3', 'Ho': 'Ho2O3', 'Er': 'Er2O3', 'Tm': 'Tm2O3', 'Yb': 'Yb2O3',
                'Lu': 'Lu2O3', 'Sb': 'Sb2O3', 'As': 'As2O5', 'Si': 'SiO2', 'Ti': 'TiO2', 'Zr': 'ZrO2', 'Hf': 'HfO2', 'Sn': 'SnO2',
                'Ge': 'GeO2', 'Te': 'TeO2', 'Se': 'SeO2', 'Th': 'ThO2', 'C': 'CO2', 'P': 'P2O5', 'V': 'V2O5', 'Nb': 'Nb2O5',
                'Ta': 'Ta2O5', 'S': 'SO3', 'Mo': 'MoO3', 'W': 'WO3', 'U': 'UO3', 'Cr6': 'CrO3'}
_OXIDES_SEEN = {'Fe2O3', 'Mn2O3', 'MnO2', 'Cu2O', 'UO2', 'PbO2', 'SnO', 'Ti2O3', 'V2O3', 'VO2', 'CrO3', 'Sb2O5', 'As2O3', 'Tl2O3',
                'CeO2', 'SO2', 'SeO3', 'TeO3', 'MoO2', 'NbO2', 'Co2O3', 'EuO', 'PdO', 'Au2O3', 'ReO2', 'Re2O7', 'Cr2O3'}

def parse_icdd_analysis(text):
    """(wt {constituent: mean wt%}, formula text, n analyses, issues) from an entry's Analysis
    field; (None, None, None, []) when it holds no wt% list followed by a formula."""
    m = re.search(r'\(wt\.?\s*%\)\s*:?\s*(.*?)(?::|;)\s*([^:;]*(?:[A-Z][a-z]?\d|\)|\]).*)$', text, re.S)
    if not m:
        return None, None, None, []
    wt_txt, formula = m.group(1), m.group(2)
    wt = OrderedDict(); issues = []
    zeros = set(re.findall(r'\b[A-Z][a-z]?\d?0\d\b', wt_txt))
    if zeros:
        issues.append('a zero written for O in the wt%% list (%s)' % ', '.join(sorted(zeros)))
        wt_txt = re.sub(r'\b([A-Z][a-z]?\d?)0(\d)\b', lambda mm: mm.group(1) + 'O' + mm.group(2), wt_txt)
    wt_txt = re.sub(r'\(\s*NH4\s*\)2\s*O', 'N2H8O', wt_txt)                # ammonium as its 'oxide'
    # an 'O=F' / 'O≡F,Cl' correction line is not a constituent (its value is the deduction)
    wt_txt = re.sub(r'-?\s*O\s*[=≡]\s*(?:F|Cl|Br|S)(?:\s*,\s*(?:F|Cl|Br|S))*\s*[-−–]?\s*\d+(?:\.\d+)?', ' ', wt_txt)
    def unglue(m):
        head, digits, frac = m.group(1), m.group(2), m.group(3)             # 'SrO' '14' '.19' / 'As2O' '560' '.33'
        el = re.match(r'[A-Z][a-z]?', head).group(0)
        best = None
        for k in range(0, min(2, len(digits) - 1) + 1):
            ox, val = head + digits[:k], digits[k:] + frac
            if not re.match(r'^\d{1,2}\.\d+$', val):
                continue
            try:
                kk = parse_constituent(ox)
            except ValueError:
                continue
            usual = _USUAL_OXIDE.get(el)
            rank = (0 if ox == usual else 1 if ox in _OXIDES_SEEN else 2, k)
            if best is None or rank < best[0]:
                best = (rank, ox, val)
        if best is None:
            return m.group(0)
        issues.append('%s: no space between the constituent and its value (read as %s %s)' % (m.group(0), best[1], best[2]))
        return '%s %s' % (best[1], best[2])
    wt_txt = re.sub(r'\b([A-Z][a-z]?\d?O)(\d{2,})(\.\d+)', unglue, wt_txt)
    for c, v in _ICDD_WT.findall(wt_txt):
        if c in ('O', 'H', 'C', 'N'):
            continue
        if c in wt:
            issues.append('%s listed twice' % c)
        try:
            wt[c] = float(v)
        except ValueError:
            pass
    for mm in re.finditer(r'\b([A-Z][a-z]?\d*O\d*|F|Cl)\s*(?:\([^)]*\))?\s*,', wt_txt):
        if mm.group(1) not in wt and mm.group(1) != 'O':
            issues.append('no value for %s' % mm.group(1))
    for c in list(wt):
        try:
            k = parse_constituent(c)
        except ValueError:
            issues.append('unknown constituent %s' % c); del wt[c]; continue
        if re.fullmatch(r'[A-Z][a-z]?O', c) and k.element not in ('Mg', 'Ca', 'Sr', 'Ba', 'Be', 'Fe', 'Mn', 'Ni', 'Co', 'Zn', 'Cu', 'Cd', 'Pb', 'Sn', 'Hg', 'Eu', 'Pd', 'Ag', 'Tl'):
            issues.append('unusual constituent %s (a digit missing?)' % c)
        elif k.kind == 'oxide' and k.element in ('Na', 'K', 'Li', 'Rb', 'Cs') and c != k.element + '2O':
            issues.append('unusual constituent %s (%s2O?)' % (c, k.element))
    tot = sum(wt.values())
    if wt and tot < 50:
        issues.append('the wt%% values add to %.1f — not a weight-per-cent list?' % tot)
    n = re.search(r'average of (\d+)', text)
    return wt, formula.strip(), (int(n.group(1)) if n else None), issues

def parse_icdd_formula(text, has_sulfur=False):
    """The ICDD formula notation -> ({element: apfu} with H and O, {element: charge}, issues).
    Handles '( Mn1.75 +2 Mg0.25 )sigma2.00' and ')S2.00' / ')$SI0.97' / ')SS1.96' sums (checked
    against the parts), '( P O4 )4' multipliers, '[ … ]7.00' sums written bare, 'OH1.09' = (OH)1.09,
    '?0.35' vacancies, 'Fe3 C1.01' (= Fe3+ 1.01), '!3.9 H2 O' hydrate water, stray brackets."""
    t = text.replace('·', ' ! ').replace('•', ' ! ').replace('□', '?')
    t = re.sub(r'([\)\]\}])\s*\$?(?:sigma|Sigma|SIGMA|SS|SI)\s*(?=\d)', r'\1 Σ', t)
    t = re.sub(r'([\)\]\}])\s*S\s*(?=\d)', r'\1 S?', t)          # a sum, or sulfur — decided by the parts
    t = re.sub(r'\s+', ' ', t)
    toks = [x for x in _ICDD_TOK.findall(t) if x.strip()]
    counts = {}; ox = {}; sums = []; items = [0.0]; issues = []
    def add(el, n, mult):
        counts[el] = counts.get(el, 0.0) + n * mult
    def parse(i, mult):
        while i < len(toks):
            tk = toks[i]
            if tk in ('(', '[', '{'):
                j = i + 1; depth = 1
                while j < len(toks) and depth:
                    if toks[j] in ('(', '[', '{'): depth += 1
                    elif toks[j] in (')', ']', '}'): depth -= 1
                    j += 1
                # the group's item total first (a dry run), to tell a sum from a multiplier
                items.append(0.0); saved = (dict(counts), dict(ox), list(sums), list(issues))
                inner(i + 1, j - 1, 1.0)
                total = items.pop(); counts.clear(); counts.update(saved[0]); ox.clear(); ox.update(saved[1]); sums[:] = saved[2]; issues[:] = saved[3]
                gm = 1.0; k = j; stated = None
                if k < len(toks) and toks[k] in ('Σ', 'sigma', 'Sigma', 'SIGMA'):
                    k += 1
                    if k < len(toks) and re.match(r'^\d', toks[k]):
                        stated = float(toks[k]); k += 1
                elif k < len(toks) and toks[k] == 'S?':
                    nxt = toks[k + 1] if k + 1 < len(toks) and re.match(r'^\d', toks[k + 1]) else None
                    if nxt is not None and abs(float(nxt) - total) <= max(0.06, 0.1 * total):
                        stated = float(nxt); k += 2                             # ')S2.000': the sum of the parts
                    elif nxt is not None and '.' not in nxt and float(nxt) <= 12:
                        gm = float(nxt); k += 2                                 # '( U O2 )S2': a subscript multiplier
                    elif nxt is not None and not has_sulfur:
                        stated = float(nxt); k += 2
                    else:
                        toks[k] = 'S'
                elif k < len(toks) and re.match(r'^\d+(\.\d+)?$', toks[k]):
                    v = float(toks[k])
                    if '.' in toks[k] and abs(v - total) <= max(0.06, 0.1 * total):
                        stated = v; k += 1                                  # ']7.00' after O5.91 OH1.09
                    else:
                        gm = v; k += 1
                items.append(0.0)
                inner(i + 1, j - 1, mult * gm)
                got = items.pop(); items[-1] += gm
                if stated is not None:
                    sums.append((stated, got))
                i = k; continue
            if tk in (')', ']', '}'):
                if len(items) == 1:
                    issues.append('unbalanced bracket in the formula'); i += 1
                    if i < len(toks) and toks[i] in ('S?', 'Σ') and i + 1 < len(toks) and re.match(r'^\d', toks[i + 1]):
                        i += 2
                    continue
                return i + 1
            if tk == '!':
                n = 1.0; i += 1
                if i < len(toks) and re.match(r'^\d+(\.\d+)?$', toks[i]):
                    n = float(toks[i]); i += 1
                if i + 2 < len(toks) and toks[i] == 'H' and toks[i + 1] == '2' and toks[i + 2] == 'O':
                    add('H', 2, n * mult); add('O', 1, n * mult); i += 3
                elif i + 1 < len(toks) and toks[i] == 'H' and toks[i + 1] == 'O':
                    add('H', 1, n * mult); add('O', 1, n * mult); i += 2
                continue
            if tk in ('?', 'S?'):
                if tk == 'S?':
                    toks[i] = 'S'; continue
                n = 1.0; i += 1
                if i < len(toks) and re.match(r'^\d+(\.\d+)?$', toks[i]):
                    n = float(toks[i]); i += 1
                items[-1] += n; continue                                      # a vacancy: an item, no atom
            if tk == 'O' and i + 2 < len(toks) and toks[i + 1] == 'H' and re.match(r'^\d+(\.\d+)?$', toks[i + 2]):
                n = float(toks[i + 2]); add('O', n, mult); add('H', n, mult); items[-1] += n; i += 3; continue
            if re.match(r'^[A-Z][a-z]?$', tk) and tk not in ATOMIC_WEIGHTS:
                if 'unrecognised symbol %s in the formula' % tk not in issues:
                    issues.append('unrecognised symbol %s in the formula' % tk)
                i += 1
                if i < len(toks) and re.match(r'^\d+(\.\d+)?$', toks[i]):
                    i += 1
                continue
            if re.match(r'^[A-Z][a-z]?$', tk):
                el = tk; n = 1.0; i += 1
                if i < len(toks) and re.match(r'^\d+(\.\d+)?$', toks[i]):
                    n = float(toks[i]); i += 1
                    if i + 1 < len(toks) and toks[i] == 'C' and re.match(r'^\d+(\.\d+)?$', toks[i + 1]) and n in (1, 2, 3, 4, 5, 6, 7) and el != 'C':
                        ox[el] = int(n); n = float(toks[i + 1]); i += 2  # 'Fe3 C1.01' = Fe3+ 1.01
                if i < len(toks) and re.match(r'^[+-]\d+(\.\d+)?$', toks[i]):
                    mm = re.match(r'^([+-]\d)(\d+\.\d+)$', toks[i])
                    if mm and n == 1.0:                                       # 'Fe +30.22': the charge and the count glued
                        ox[el] = float(mm.group(1)); n = float(mm.group(2))
                    else:
                        ox[el] = float(toks[i])
                    i += 1
                add(el, n, mult); items[-1] += n; continue
            i += 1
        return i
    def inner(a, b, mult):
        sub = toks[a:b]; saved = toks[:]
        toks[:] = sub
        parse(0, mult)
        toks[:] = saved
    parse(0, 1.0)
    bad = [(st_, got) for st_, got in sums if abs(st_ - got) > max(0.03, 0.03 * st_)]
    if bad:
        issues.append('formula group sums do not add up: ' + ', '.join('Σ%.2f given, parts add to %.2f' % x for x in bad[:3]))
    big = [el for el, v in counts.items() if el not in ('H', 'O') and v > 40]
    if big:
        issues.append('formula group sums do not add up: %s apfu is not a coefficient (%s)' % (', '.join('%s %.1f' % (el, counts[el]) for el in big), 'a garbled multiplier or sum'))
    return counts, ox, issues

def basis_candidates(counts):
    """The bases a paper may have used, read off the formula: the anion total (with and without
    the water O), the cation total when it is a whole number, and every whole-number element."""
    cands = []
    o_tot = counts.get('O', 0.0) + counts.get('F', 0.0) + counts.get('Cl', 0.0)
    h2o = counts.get('H', 0.0) / 2
    for n in sorted({round(o_tot), round(o_tot - h2o), round(o_tot - h2o - counts.get('F', 0.0))}, reverse=True):
        if n > 0:
            cands.append(('O', float(n)))
    cats = {k: v for k, v in counts.items() if k not in ('H', 'O', 'F', 'Cl', 'Br', 'I', 'S', 'Se', 'Te')}
    tot = sum(cats.values())
    if tot > 0 and abs(tot - round(tot)) < 0.08:
        cands.append(('cations', float(round(tot))))
    chalc = sum(counts.get(k, 0.0) for k in ('S', 'Se', 'Te'))
    if chalc >= 0.9 and not counts.get('O'):
        cands.append(('element', '+'.join(k for k in ('S', 'Se', 'Te') if counts.get(k)), float(round(chalc))))
    for el, v in cats.items():
        if v >= 0.9 and abs(v - round(v)) < 0.06:
            cands.append(('element', el, float(round(v))))
    if counts.get('S', 0) >= 0.9 and counts.get('O'):
        cands.append(('element', 'S', float(round(counts['S']))))
    return cands

def _apfu_of(red):
    out = {}
    for k, r in red.rows.items():
        el = 'H' if r.c.kind == 'water' else r.c.element
        out[el] = out.get(el, 0.0) + r.apfu
        if r.c.formula == 'N2H8O':                          # ammonium: its H count too
            out['H'] = out.get('H', 0.0) + 4 * r.apfu
    return out

def replicate_formula(wt, counts, bases, name='entry', tol_abs=0.02, tol_rel=0.02):
    """Re-reduce the mean wt% on each candidate basis and compare with the published apfu.
    -> {'basis', 'score' (rms relative deviation of the cations), 'apfu', 'reduction', 'diffs':
    [(element, published, replicated, note)], 'unanalysed': [elements the probe did not give]}."""
    cons, vals = [], []
    for c, v in wt.items():
        try:
            cons.append(parse_constituent(c)); vals.append(v)
        except ValueError:
            pass
    ds = Dataset(cons, [vals], ['mean'], {}, name, None)
    best = None
    for b in bases:
        try:
            red = reduce(ds, b)
        except Exception:
            continue
        apfu = _apfu_of(red)
        devs = [(apfu[el] - v) / max(v, 0.05) for el, v in counts.items() if el not in ('H', 'O') and v >= 0.05 and el in apfu]
        if not devs:
            continue
        sc = math.sqrt(sum(d * d for d in devs) / len(devs))
        if best is None or sc < best[0]:
            best = (sc, b, apfu, red)
    if best is None:
        return None
    sc, b, apfu, red = best
    diffs = []; unanalysed = []
    # a constant factor between every replicated and published cation means the paper's basis
    # was not among the candidates — the numbers themselves agree
    ratios = [apfu[el] / v for el, v in counts.items() if el not in ('H', 'O', 'F', 'Cl') and v >= 0.2 and el in apfu and apfu[el] > 0]
    factor = None
    if len(ratios) >= 2:
        med = sorted(ratios)[len(ratios) // 2]
        if abs(med - 1) > 0.02 and all(abs(r / med - 1) <= 0.02 for r in ratios):
            factor = med
    counts = dict(counts)
    if counts.get('N') and 'N' not in apfu:
        counts['H'] = counts.get('H', 0.0) - 4 * counts['N']                       # ammonium H: not from the H2O
    for el, v in counts.items():
        if el == 'O':
            continue
        t = apfu.get(el)
        if factor and t is not None:
            t = t / factor
        if t is None:
            if v >= 0.05:
                (unanalysed if el in CALCULATED_ELEMENTS else diffs).append((el, v, None, 'not in the wt% list'))
            continue
        d = t - v
        if abs(d) > max(tol_abs, tol_rel * v):
            diffs.append((el, v, t, '%+.3f' % d))
    return {'basis': b, 'score': sc, 'apfu': apfu, 'reduction': red, 'diffs': diffs, 'unanalysed': [u[0] for u in unanalysed], 'factor': factor}

def check_analysis(text, basis=None, name='entry'):
    """One entry's Analysis field: parse, replicate, report. -> (lines, result dict or None).
    `basis` ('O=21', 'cations=8', 'Si+Al=4') forces the basis; otherwise the candidates read off
    the formula are tried and the one that reproduces it best is reported."""
    wt, ftxt, n, issues = parse_icdd_analysis(text)
    L = []
    if not wt or len(wt) < 2 or not ftxt:
        return ['no wt% list followed by a formula in the Analysis field'], None
    has_s = any(c in ('S', 'SO3', 'SO2') for c in wt)
    counts, ox, f_issues = parse_icdd_formula(ftxt, has_s)
    issues += f_issues
    if not re.search(r'\d\.\d', ftxt):
        return ['the formula given is an ideal one (no decimals) — nothing to replicate'] + ['! ' + x for x in issues], None
    if any('group sums' in x or 'unbalanced' in x for x in issues):
        return ['formula notation problem: ' + '; '.join(issues)], None
    bases = [_parse_basis(basis)] if basis else basis_candidates(counts)
    r = replicate_formula(wt, counts, bases, name)
    if r is None:
        return ['no basis reproduces the formula (%s)' % ', '.join(_basis_label(b) for b in bases)] + ['! ' + x for x in issues], None
    r['issues'] = issues; r['wt'] = wt; r['published'] = counts; r['n'] = n
    red = r['reduction']
    L.append('basis %s%s; rms deviation of the cations %.1f%%; total %.2f wt%%' % (_basis_label(r['basis']), '' if basis else ' (inferred from the formula)', 100 * r['score'], red.total))
    if r.get('factor'):
        L.append('  the published coefficients are the replicated ones ÷ %.3f — the paper normalised on another basis; the numbers agree' % r['factor'])
    L.append('replicated: ' + red.formula())
    for el, v, t, note in r['diffs']:
        L.append('  %-3s published %.3f  replicated %s  %s' % (el, v, ('%.3f' % t) if t is not None else '—', note))
    if r['unanalysed']:
        L.append('  (calculated by the authors, not analysed: %s)' % ', '.join(r['unanalysed']))
    for x in issues:
        L.append('  ! ' + x)
    if not r['diffs'] and not issues:
        L.append('  every coefficient follows from the wt%% list within %s' % ('0.02 apfu / 2 %'))
    return L, r

# ----------------------------------------------------------------------------- CLI

def _parse_basis(s):
    s = (s or 'O=1').strip()
    m = re.fullmatch(r'(O|anions)\s*=\s*([\d.]+)', s, re.I)
    if m:
        return ('O', float(m.group(2)))
    m = re.fullmatch(r'cations?\s*=\s*([\d.]+)', s, re.I)
    if m:
        return ('cations', float(m.group(1)))
    m = re.fullmatch(r'([A-Za-z+ ]+)\s*=\s*([\d.]+)', s)
    if m:
        return ('element', m.group(1).replace(' ', ''), float(m.group(2)))
    raise ValueError('basis must be O=N, cations=N or El=N (e.g. U=5, Si+Al=4): %r' % s)

def _parse_add(s):
    m = re.fullmatch(r'([A-Za-z0-9]+)\s*=\s*(structure|wt)\s*:\s*([\d.]+)|([A-Za-z0-9]+)\s*=\s*(difference)', s.strip())
    if not m:
        raise ValueError('an addition must be X=structure:N, X=wt:V or X=difference: %r' % s)
    if m.group(5):
        return (m.group(4), 'difference', None)
    return (m.group(1), m.group(2), float(m.group(3)))

def parse_points(s):
    """'1-8' / '1,3,5' (1-based) -> 0-based indices; None for blank."""
    if not (s or '').strip():
        return None
    pts = []
    for tok in s.split(','):
        tok = tok.strip()
        if not tok:
            continue
        m = re.fullmatch(r'(\d+)\s*-\s*(\d+)', tok)
        if m:
            pts += list(range(int(m.group(1)) - 1, int(m.group(2))))
        elif tok.isdigit():
            pts.append(int(tok) - 1)
        else:
            raise ValueError('points must be like 1-8 or 1,3,5: %r' % s)
    return pts

def _parse_pairs(s):
    out = OrderedDict()
    for tok in (s or '').split(','):
        if '=' in tok:
            k, v = tok.split('=', 1); out[k.strip()] = v.strip()
    return out

def prepare(probe, basis='O=1', charge=None, anions=None, raw_anions=False, adds=(), converts=(), drop=(),
            points=None, standards=None, ideal=None, name='', sheet=None, method=''):
    """The whole reduction from CLI-style strings — shared by main() and the GUI. Returns
    (dataset, reduction, published table, report text); ValueError on any bad input."""
    ds = load_probe(probe, sheet or None)
    pts = parse_points(points)
    adds_ = [_parse_add(a) for a in adds if a.strip()]
    conv = []
    for c in converts:
        if c.strip():
            if '=' not in c:
                raise ValueError('a conversion must be like UO2=UO3: %r' % c)
            conv.append(tuple(x.strip() for x in c.split('=', 1)))
    drop = [d.strip() for d in drop if d.strip()]
    b = _parse_basis(basis)
    red = charge_balance(ds, b, adds_, conv, drop, pts, charge, anions, raw_anions) if charge \
        else reduce(ds, b, adds_, conv, drop, pts, raw_anions=raw_anions)
    stds = _parse_pairs(standards)
    ideal_ = None
    if ideal and ideal.strip():
        try:
            ideal_ = ideal_wt_percent(OrderedDict((k, float(v)) for k, v in _parse_pairs(ideal).items()))
        except ValueError:
            raise ValueError('the ideal formula must be like CaO=3,UO3=1,H2O=6: %r' % ideal)
    table = published_table(red, stds, ideal_, name)
    return ds, red, table, report_text(red, table)

def to_tabs(table, journal_key=None):
    """The published table as a pxrd_review.tables table list (for render_html / write_word)."""
    from pxrd_review import tables as T
    J = T.journal(journal_key)
    note = table['note']
    if note and J['notes_prefix'] and not note.startswith(J['notes_prefix']):
        note = J['notes_prefix'] + note
    return [('epma', {'n': 1, 'label': J['caption'].format(n=1), 'caption': T._title(J, table['caption'].rstrip('.')), 'journal': J,
                      'head': [T.C(h) for h in table['head']], 'rows': [[T.C(str(x)) for x in r] for r in table['rows']], 'note': note})]

def export(ds, red, table, text, out_dir, stem, word=False, xlsx=False, journal_key=None, method=''):
    """Write review_out/<stem>_epma.txt (always) and the .xlsx / .docx asked for; {kind: path}."""
    os.makedirs(out_dir, exist_ok=True)
    paths = {}
    with open(os.path.join(out_dir, stem + '_epma.txt'), 'w', encoding='utf-8') as f:
        f.write(text + '\n')
    paths['text'] = os.path.join(out_dir, stem + '_epma.txt')
    if xlsx:
        paths['xlsx'] = write_xlsx(red, table, os.path.join(out_dir, stem + '_epma.xlsx'), ds, method)
    if word:
        from pxrd_review import tables as T
        paths['word'] = os.path.join(out_dir, stem + '_epma.docx'); T.write_word(None, to_tabs(table, journal_key), paths['word'])
    return paths

def main(argv=None):
    ap = argparse.ArgumentParser(prog='pxrd epma', description=__doc__.split('\n\n')[1], formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('probe', nargs='?', help='the probe export (xlsx / csv / txt) — or, with --check, an ICDD entry .docx')
    ap.add_argument('--check', metavar='TEXT', nargs='?', const='', help='replicate a published formula from an ICDD Analysis string '
                    '("Microprobe analysis … (wt.%%): CaO 18.4, … : Ca2.99Na0.67…"); with an entry .docx as the positional argument, its Analysis field')
    ap.add_argument('--basis', default='O=1', help='O=N (anions), cations=N, or El=N / El+El=N')
    ap.add_argument('--charge', choices=['H2O', 'Fe'], help='balance charge: H2O = add the hydrogen the analysis lacks (needs a cation/element basis and --anions N); Fe = split FeO into FeO + Fe2O3 on the basis given')
    ap.add_argument('--anions', type=float, help='total anions (O + F) per formula unit from the structure, for --charge H2O')
    ap.add_argument('--raw-anions', action='store_true', help='count oxide O and halogens both, without the O=F reduction (the spreadsheet convention, e.g. 21.5)')
    ap.add_argument('--add', action='append', default=[], help='X=structure:N | X=wt:V | X=difference (repeatable)')
    ap.add_argument('--convert', action='append', default=[], help='UO2=UO3 (repeatable)')
    ap.add_argument('--drop', action='append', default=[], help='constituent to leave out (repeatable)')
    ap.add_argument('--points', help='rows to use, e.g. 1-8 or 1,3,5 (1-based, default all)')
    ap.add_argument('--standards', help='CaO=anorthite,UO3=UO2 …')
    ap.add_argument('--ideal', help='ideal apfu as X=N,… e.g. CaO=3,Na2O=1,UO3=1,SeO2=1,CO2=3,F=1,H2O=6')
    ap.add_argument('--name', default='', help='mineral name for the caption')
    ap.add_argument('--sheet'); ap.add_argument('--out'); ap.add_argument('--xlsx', action='store_true'); ap.add_argument('--word', action='store_true')
    ap.add_argument('--journal', default='manuscript')
    a = ap.parse_args(argv)
    if a.check is not None:
        text = a.check
        if not text and a.probe and a.probe.lower().endswith('.docx'):
            from pxrd_review.extra_checks import parse_entry
            text = (parse_entry(a.probe).comments.get('Analysis') or '').strip()
        if not text:
            raise SystemExit('epma --check: give the Analysis string, or an entry .docx')
        lines, r = check_analysis(text, None if a.basis == 'O=1' else a.basis, os.path.basename(a.probe or 'entry'))
        print('\n'.join(lines))
        return 0 if r is not None and not r['diffs'] and not r['issues'] else 1
    if not a.probe:
        raise SystemExit('epma: give the probe export, e.g.  pxrd epma probe.xlsx --basis O=21 --xlsx')
    try:
        ds, red, table, text = prepare(a.probe, a.basis, a.charge, a.anions, a.raw_anions, a.add, a.convert, a.drop,
                                       a.points, a.standards, a.ideal, a.name, a.sheet)
    except ValueError as e:
        raise SystemExit('epma: %s' % e)
    print(text)
    out = a.out or os.path.join(os.path.dirname(os.path.abspath(a.probe)), 'review_out')
    stem = os.path.splitext(os.path.basename(a.probe))[0]
    paths = export(ds, red, table, text, out, stem, a.word, a.xlsx, a.journal)
    for k in ('xlsx', 'word'):
        if k in paths:
            print('  %s → %s' % (k, paths[k]))
    return 0

if __name__ == '__main__':
    sys.exit(main())
