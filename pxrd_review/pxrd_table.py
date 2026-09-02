#!/usr/bin/env python3
"""
pxrd_table — the powder X-ray diffraction table of a mineral description: observed lines beside the
calculated ones, overlapping calculated reflections combined under one observed line, the eight
strongest observed lines in bold, two column blocks.

    python3 -m pxrd_review.pxrd_table <obs list> <calc list> [--min-icalc 3.5] [--tol 1.0] [--name X]
                                      [--journal ammin] [--word] [--xlsx] [--out DIR] [--wavelength 1.5406]
    pxrd pxrd obs.txt calc.txt --word

Inputs: JADE peak lists (the observed list from a whole-pattern fit — with the hkl JADE assigned —
and the calculated pattern of the structure), or any text/csv with a header naming the columns
(d, I, hkl, 2θ/Angle), or a bare two-column d/I (or 2θ/I with --wavelength) list.

Matching: an observed line and a calculated one are the same reflection when their hkl agree;
observed lines without an hkl take the nearest calculated line within --tol % in d. An observed
peak that JADE resolved into several reflections appears once per reflection with its Iobs/dobs
repeated (the usual layout); calculated lines nobody observed are listed with blank Iobs/dobs when
Icalc ≥ --min-icalc. Intensities are printed as integers, d to three decimals; the eight strongest
observed peaks are bold (CNMNC practice).
"""
import os, re, sys, math, argparse
from collections import namedtuple, OrderedDict

Line = namedtuple('Line', 'd I hkl two_theta extra')

_HKL = re.compile(r'\(\s*(-?\d+)\s+(-?\d+)\s+(-?\d+)\s*\)')

def _hkl_of(tok):
    m = _HKL.search(tok or '')
    return (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None

def _split(line):
    if '\t' in line:
        return [c.strip() for c in line.split('\t')]
    # keep '( 1 0 0)' together
    parts = re.split(r'\s{2,}|\s(?=\S)', line.strip())
    out = []; buf = ''
    for p in parts:
        if buf:
            buf += ' ' + p
            if ')' in p:
                out.append(buf); buf = ''
        elif p.startswith('(') and ')' not in p:
            buf = p
        else:
            out.append(p)
    if buf:
        out.append(buf)
    return out

def _num(s):
    try:
        return float(str(s).replace(',', '.'))
    except (TypeError, ValueError):
        return None

def load_lines(path, wavelength=None):
    """[Line] from a JADE list or any headed / bare list."""
    with open(path, encoding='utf-8-sig', errors='replace') as f:
        rows = [ln.rstrip('\n') for ln in f]
    header = None; cols = {}; out = []
    for ln in rows:
        if not ln.strip():
            continue
        cells = _split(ln)
        low = [c.lower() for c in cells]
        if header is None and any(re.match(r'd\s*\(|^d$|d-?spac|dobs|dcalc', c) for c in low) or \
                (header is None and any(c.startswith(('i%', 'int', 'i(', 'iobs', 'icalc')) or c == 'i' for c in low)):
            header = cells
            for i, c in enumerate(low):
                if re.match(r'd\s*\(|^d$|d-?spac|dobs|dcalc', c) and 'd' not in cols:
                    cols['d'] = i
                elif (c.startswith(('i%', 'int', 'i(', 'iobs', 'icalc')) or c == 'i') and 'I' not in cols:
                    cols['I'] = i
                elif ('h k l' in c or c.startswith('hkl') or c == '( h k l)') and 'hkl' not in cols:
                    cols['hkl'] = i
                elif (c.startswith('angle') or '2θ' in c or '2theta' in c or c.startswith('2t')) and 'tt' not in cols:
                    cols['tt'] = i
            continue
        nums = [_num(c) for c in cells]
        if header is not None and cols:
            try:
                d = _num(cells[cols['d']]) if 'd' in cols else None
                I = _num(cells[cols['I']]) if 'I' in cols else None
                tt = _num(cells[cols['tt']]) if 'tt' in cols else None
                hkl = _hkl_of(cells[cols['hkl']]) if 'hkl' in cols else None
            except IndexError:
                continue
            if hkl is None:
                hkl = next((h for h in (_hkl_of(c) for c in cells) if h), None)
            if d is None and tt is not None and wavelength:
                d = wavelength / (2 * math.sin(math.radians(tt / 2)))
            if d is None or I is None:
                continue
            out.append(Line(d, I, hkl, tt, cells))
        else:
            # bare list: the first two numbers are d (or 2θ) and I; an hkl triple if present
            vals = [v for v in nums if v is not None]
            if len(vals) < 2:
                continue
            hkl = next((h for h in (_hkl_of(c) for c in cells) if h), None)
            if hkl is None and len(vals) >= 5 and all(float(v).is_integer() for v in vals[2:5]):
                hkl = tuple(int(v) for v in vals[2:5])
            a, I = vals[0], vals[1]
            d = wavelength / (2 * math.sin(math.radians(a / 2))) if (wavelength and a > 20) else a
            out.append(Line(d, I, hkl, a if d != a else None, cells))
    if not out:
        raise ValueError('no peaks read from %s' % path)
    return out

# ----------------------------------------------------------------------------- matching

Row = namedtuple('Row', 'Iobs dobs dcalc Icalc hkl obs_id')

def match(obs, calc, tol_pct=1.2, min_i=3.5, dmin=None):
    """-> [Row] in order of decreasing d.
    1. an observed line with an hkl takes the calculated line of that hkl; without one, the nearest
       calculated line within tol_pct in d;
    2. a calculated line nobody claimed is ATTACHED to the nearest observed peak within tol_pct
       (its Iobs/dobs repeated — the overlap the observed peak really contains), else listed alone;
    3. a row survives when Iobs or Icalc is >= min_i; a kept observed line whose calculated match is
       weaker than min_i shows its hkl with dcalc/Icalc blank.
    `obs_id` groups the rows of one observed peak."""
    by_hkl = {}
    for i, c in enumerate(calc):
        if c.hkl:
            by_hkl.setdefault(c.hkl, i)
    used = set(); rows = []
    peaks = OrderedDict()
    for o in obs:
        peaks.setdefault((round(o.d, 4), round(o.I, 2)), []).append(o)
    peak_d = {}
    for pid, (key, group) in enumerate(peaks.items()):
        peak_d[pid] = group[0]
        for o in group:
            ci = by_hkl.get(o.hkl) if o.hkl else None
            if ci is None:
                best = None
                for i, c in enumerate(calc):
                    if i in used:
                        continue
                    if abs(c.d - o.d) / o.d * 100 <= tol_pct and (best is None or abs(c.d - o.d) < abs(calc[best].d - o.d)):
                        best = i
                ci = best
            if ci is None:
                rows.append(Row(o.I, o.d, None, None, o.hkl, pid)); continue
            used.add(ci); c = calc[ci]
            rows.append(Row(o.I, o.d, c.d, c.I, c.hkl or o.hkl, pid))
    # unclaimed calculated lines: attach to the nearest observed peak (one strong enough to be
    # listed) within tol, else stand alone
    for i, c in enumerate(calc):
        if i in used:
            continue
        best = None
        for pid, o in peak_d.items():
            if o.I < min_i:
                continue
            if abs(c.d - o.d) / o.d * 100 <= tol_pct and (best is None or abs(c.d - o.d) < abs(peak_d[best].d - c.d)):
                best = pid
        if best is not None:
            o = peak_d[best]
            rows.append(Row(o.I, o.d, c.d, c.I, c.hkl, best))
        else:
            rows.append(Row(None, None, c.d, c.I, c.hkl, None))
    # thresholds: a row lives when Iobs or Icalc reaches min_i; an observed peak's extra reflections
    # whose calculated line is weak are dropped when the peak already has a real match, and kept
    # (hkl shown, calc blank) only when they are all the peak has
    strong = {}
    for r in rows:
        if r.obs_id is not None and r.Icalc is not None and r.Icalc >= min_i:
            strong[r.obs_id] = True
    # a peak with only weak reflections keeps ONE row: its strongest reflection, calc blank
    weak_best = {}
    for r in rows:
        if r.obs_id is not None and not strong.get(r.obs_id):
            ic = r.Icalc if r.Icalc is not None else -1
            if r.obs_id not in weak_best or ic > weak_best[r.obs_id][0]:
                weak_best[r.obs_id] = (ic, r.hkl)
    kept = []
    for r in rows:
        io = r.Iobs if r.Iobs is not None else 0.0; ic = r.Icalc if r.Icalc is not None else 0.0
        if max(io, ic) < min_i:
            continue
        if r.Iobs is not None and (r.Icalc is None or r.Icalc < min_i):
            if strong.get(r.obs_id) or weak_best[r.obs_id][1] != r.hkl:
                continue
            r = r._replace(dcalc=None, Icalc=None)
        kept.append(r)
    if dmin:
        kept = [r for r in kept if (r.dobs if r.dobs is not None else r.dcalc) >= dmin]
    kept.sort(key=lambda r: (-(r.dobs if r.dobs is not None else r.dcalc), r.obs_id if r.obs_id is not None else 10**9,
                             -(r.dcalc if r.dcalc is not None else 0)))
    return kept

def strongest(rows, n=8):
    """obs_ids of the n strongest observed peaks."""
    seen = {}
    for r in rows:
        if r.obs_id is not None and r.Iobs is not None:
            seen[r.obs_id] = r.Iobs
    return set(k for k, _ in sorted(seen.items(), key=lambda kv: -kv[1])[:n])

# ----------------------------------------------------------------------------- table

def _hkl_str(h):
    return ' '.join(str(x) for x in h) if h else ''

def build_table(rows, name='', journal_key=None, min_i=3.5, blocks=2, bold_n=8):
    from pxrd_review import tables as T
    J = T.journal(journal_key)
    top = strongest(rows, bold_n)
    cells = []
    for r in rows:
        b = r.obs_id in top
        iobs = T.R('%d' % round(r.Iobs), *(['b'] if b else [])) if r.Iobs is not None else T.C('')
        dobs = T.R('%.3f' % r.dobs, *(['b'] if b else [])) if r.dobs is not None else T.C('')
        cells.append([iobs, dobs, T.C('%.3f' % r.dcalc if r.dcalc is not None else ''),
                      T.C('%d' % round(r.Icalc) if r.Icalc is not None else ''), T.C(_hkl_str(r.hkl))])
    head1 = [T.C(T.R('I', 'i'), T.R('obs', 'sub')), T.C(T.R('d', 'i'), T.R('obs', 'sub')), T.C(T.R('d', 'i'), T.R('calc', 'sub')),
             T.C(T.R('I', 'i'), T.R('calc', 'sub')), T.C(T.R('hkl', 'i'))]
    n = len(cells); per = math.ceil(n / blocks) if n else 0
    out_rows = []
    for i in range(per):
        row = []
        for b in range(blocks):
            j = b * per + i
            if b:
                row.append(T.C(''))
            row += cells[j] if j < n else [T.C('')] * 5
        out_rows.append(row)
    head = []
    for b in range(blocks):
        if b:
            head.append(T.C(''))
        head += head1
    cap = 'Powder X-ray diffraction data (%s in Å) for %s' % ('d', name or '…')
    note = 'Only lines with I ≥ %g (observed or calculated) are listed; the eight strongest observed lines are in bold.' % min_i
    return {'n': 1, 'label': J['caption'].format(n=1), 'caption': T._title(J, cap), 'head': head, 'rows': out_rows,
            'note': (J['notes_prefix'] + note) if J['notes_prefix'] else note, 'journal': J}

def write_xlsx(obs, calc, rows, path):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'matched'
    ws.append(['Iobs', 'dobs', 'dcalc', 'Icalc', 'h', 'k', 'l', 'Δd (%)'])
    top = strongest(rows)
    for r in rows:
        ws.append([r.Iobs, r.dobs, r.dcalc, r.Icalc] + (list(r.hkl) if r.hkl else [None] * 3)
                  + [(r.dobs - r.dcalc) / r.dobs * 100 if r.dobs and r.dcalc else None])
        if r.obs_id in top:
            for c in (1, 2):
                ws.cell(ws.max_row, c).font = Font(bold=True)
    for title, lines in (('obs', obs), ('calc', calc)):
        w2 = wb.create_sheet(title); w2.append(['d', 'I', 'h', 'k', 'l', '2θ'])
        for l in lines:
            w2.append([l.d, l.I] + (list(l.hkl) if l.hkl else [None] * 3) + [l.two_theta])
    wb.save(path); return path

def prepare(obs_path, calc_path, tol_pct=1.2, min_i=3.5, dmin=None, wavelength=None, name='', journal_key=None, blocks=2):
    """Read both lists, match, build the table — shared by main() and the GUI; ValueError on bad input."""
    obs = load_lines(obs_path, wavelength); calc = load_lines(calc_path, wavelength)
    if not obs or not calc:
        raise ValueError('no lines read from %s' % ('the observed list' if not obs else 'the calculated list'))
    rows = match(obs, calc, tol_pct, min_i, dmin)
    return obs, calc, rows, build_table(rows, name, journal_key, min_i, max(1, min(4, int(blocks or 2))))

def export(obs, calc, rows, table, out_dir, stem, word=False, xlsx=False):
    """review_out/<stem>_pxrd.txt (always) and the .docx / .xlsx asked for; {kind: path}."""
    from pxrd_review import tables as T
    os.makedirs(out_dir, exist_ok=True)
    paths = {'text': os.path.join(out_dir, stem + '_pxrd.txt')}
    with open(paths['text'], 'w', encoding='utf-8') as f:
        f.write(T.render_text([('pxrd', table)]))
    if word:
        paths['word'] = os.path.join(out_dir, stem + '_pxrd.docx'); T.write_word(None, [('pxrd', table)], paths['word'])
    if xlsx:
        paths['xlsx'] = write_xlsx(obs, calc, rows, os.path.join(out_dir, stem + '_pxrd.xlsx'))
    return paths

def main(argv=None):
    ap = argparse.ArgumentParser(prog='pxrd pxrd', description=__doc__.split('\n\n')[1], formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('obs'); ap.add_argument('calc')
    ap.add_argument('--min-i', type=float, default=3.5, help='a line is listed when Iobs or Icalc ≥ this (default 3.5)')
    ap.add_argument('--tol', type=float, default=1.2, help='d tolerance (%%) to attach a calculated line to an observed peak (default 1.2)')
    ap.add_argument('--wavelength', type=float, help='λ (Å) to convert a 2θ list to d')
    ap.add_argument('--dmin', type=float, help='drop lines below this d (Å) — the table\'s 2θ limit')
    ap.add_argument('--name', default=''); ap.add_argument('--journal', default='manuscript'); ap.add_argument('--blocks', type=int, default=2)
    ap.add_argument('--word', action='store_true'); ap.add_argument('--xlsx', action='store_true'); ap.add_argument('--out')
    a = ap.parse_args(argv)
    try:
        obs, calc, rows, t = prepare(a.obs, a.calc, a.tol, a.min_i, a.dmin, a.wavelength, a.name, a.journal, a.blocks)
    except ValueError as e:
        raise SystemExit('pxrd: %s' % e)
    from pxrd_review import tables as T
    print(T.render_text([('pxrd', t)]))
    print('  %d observed lines, %d calculated; %d rows (%d calc-only)' % (len(obs), len(calc), len(rows), sum(1 for r in rows if r.Iobs is None)))
    out = a.out or os.path.join(os.path.dirname(os.path.abspath(a.obs)), 'review_out')
    stem = (a.name or os.path.splitext(os.path.basename(a.obs))[0]).replace(' ', '_')
    paths = export(obs, calc, rows, t, out, stem, a.word, a.xlsx)
    for k in ('word', 'xlsx'):
        if k in paths:
            print('  %s → %s' % (k, paths[k]))
    return 0

if __name__ == '__main__':
    sys.exit(main())
