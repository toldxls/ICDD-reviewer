"""Unit tests for pxrd_review.epma — probe reduction on textbook compositions.

    python3 -m unittest tests.test_epma -v
"""
import os, shutil, tempfile, unittest, csv

from pxrd_review import epma as E


def _csv(tmp, name, heads, rows):
    p = os.path.join(tmp, name)
    with open(p, 'w', newline='') as f:
        w = csv.writer(f); w.writerow(['point'] + heads)
        for i, r in enumerate(rows, 1):
            w.writerow([i] + r)
    return p


class Constituents(unittest.TestCase):
    def test_parse(self):
        c = E.parse_constituent('Al2O3'); self.assertEqual((c.element, c.n_cat, c.n_o, c.charge, c.kind), ('Al', 2, 3, 3, 'oxide'))
        self.assertAlmostEqual(c.mw, 101.96, places=1)
        self.assertEqual(E.parse_constituent('UO3').charge, 6)
        self.assertEqual(E.parse_constituent('F').kind, 'element-anion')
        self.assertEqual(E.parse_constituent('H2O').kind, 'water')
        self.assertEqual(E.parse_constituent('CO2*').formula, 'CO2')
        with self.assertRaises(ValueError):
            E.parse_constituent('Xx2O3')


class Reduction(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp(prefix='epma_')

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_anorthite_fixed_oxygen(self):
        # CaAl2Si2O8: CaO 20.16, Al2O3 36.65, SiO2 43.19 wt%
        p = _csv(self.tmp, 'an.csv', ['CaO', 'Al2O3', 'SiO2'], [[20.16, 36.65, 43.19], [20.20, 36.60, 43.20]])
        ds = E.load_probe(p)
        self.assertEqual([c.formula for c in ds.constituents], ['CaO', 'Al2O3', 'SiO2'])
        self.assertEqual(len(ds.points), 2)
        r = E.reduce(ds, ('O', 8))
        self.assertAlmostEqual(r.rows['CaO'].apfu, 1.00, places=2)
        self.assertAlmostEqual(r.rows['Al2O3'].apfu, 2.00, places=2)
        self.assertAlmostEqual(r.rows['SiO2'].apfu, 2.00, places=2)
        self.assertAlmostEqual(r.charge, 0.0, places=2)
        self.assertEqual(r.formula(), 'CaAl2Si2O8')
        # the same on a cation basis and on Si + Al = 4
        self.assertAlmostEqual(E.reduce(ds, ('cations', 5)).rows['SiO2'].apfu, 2.00, places=2)
        self.assertAlmostEqual(E.reduce(ds, ('element', 'Si+Al', 4)).rows['CaO'].apfu, 1.00, places=2)

    def test_gypsum_water_and_charge(self):
        # CaSO4·2H2O: CaO 32.57, SO3 46.50, H2O 20.93
        p = _csv(self.tmp, 'gy.csv', ['CaO', 'SO3'], [[32.57, 46.50]])
        ds = E.load_probe(p)
        r = E.reduce(ds, ('O', 6), adds=[('H2O', 'wt', 20.93)])
        self.assertAlmostEqual(r.rows['H2O'].apfu, 4.00, places=1)               # 4 H per f.u.
        self.assertAlmostEqual(r.rows['CaO'].apfu, 1.00, places=2)
        self.assertIn('·2H2O', r.formula())
        # by difference and by charge balance both recover the water
        r2 = E.reduce(ds, ('O', 6), adds=[('H2O', 'difference', None)])
        self.assertAlmostEqual(r2.rows['H2O'].wt, 20.93, places=1)
        with self.assertRaises(ValueError):
            E.charge_balance(ds, ('O', 6), adds=[], adjust='H2O')            # water is neutral: indeterminate
        r3 = E.charge_balance(ds, ('cations', 2), adds=[], adjust='H2O', anions=6)
        self.assertAlmostEqual(r3.rows['H2O'].wt, 20.93, places=1)
        self.assertAlmostEqual(r3.rows['H2O'].apfu, 4.0, places=2)
        self.assertAlmostEqual(r3.charge, 0.0, places=3)
        # structure-based: 2 H2O apfu -> the same wt%
        r4 = E.reduce(ds, ('O', 6), adds=[('H2O', 'structure', 4.0)])   # 4 H apfu
        self.assertAlmostEqual(r4.rows['H2O'].wt, 20.93, places=1)

    def test_fluorite_correction_and_fe_split(self):
        # fluorapatite-like check: F=O correction applied to the total
        p = _csv(self.tmp, 'ap.csv', ['CaO', 'P2O5', 'F'], [[55.60, 42.22, 3.77]])
        r = E.reduce(E.load_probe(p), ('O', 13))            # 12 O + 1 F
        self.assertAlmostEqual(r.corr, 3.77 * 15.999 / 18.998 / 2, places=3)
        self.assertAlmostEqual(r.rows['F'].apfu, 1.0, places=1)
        self.assertAlmostEqual(r.rows['CaO'].apfu, 5.0, places=1)              # O=F reduction on the basis
        self.assertAlmostEqual(r.o_apfu() + r.rows['F'].apfu, 13.0, places=2)
        raw = E.reduce(E.load_probe(p), ('O', 13.5), raw_anions=True)           # the spreadsheet convention
        self.assertAlmostEqual(raw.rows['CaO'].apfu, 5.0, places=1)
        # magnetite as 'FeO' 93.1 wt% (all Fe as FeO) -> the split that balances 4 O is Fe3+ = 2/3
        p2 = _csv(self.tmp, 'mt.csv', ['FeO'], [[93.11]])
        r = E.charge_balance(E.load_probe(p2), ('cations', 3), adjust='Fe', anions=4)
        self.assertAlmostEqual(r.rows['Fe2O3'].apfu / (r.rows['Fe2O3'].apfu + r.rows['FeO'].apfu), 2 / 3, places=2)
        self.assertAlmostEqual(r.charge, 0.0, places=2)

    def test_conversion_and_table_and_xlsx(self):
        p = _csv(self.tmp, 'u.csv', ['UO2', 'CaO'], [[80.0, 10.0], [81.0, 9.0]])
        ds = E.load_probe(p)
        r = E.reduce(ds, ('O', 4), converts=[('UO2', 'UO3')])
        self.assertIn('UO3', r.rows); self.assertNotIn('UO2', r.rows)
        self.assertAlmostEqual(r.rows['UO3'].wt, 80.5 * 286.03 / 270.03, places=2)
        t = E.published_table(r, {'CaO': 'anorthite'}, E.ideal_wt_percent({'UO3': 1, 'CaO': 1}), 'test')
        self.assertEqual(t['head'][:4], ['Constituent', 'Mean', 'Range', 'S.D.'])
        self.assertEqual(t['rows'][-1][0], 'Total')
        self.assertEqual(t['rows'][1][4], 'anorthite')
        out = os.path.join(self.tmp, 'out.xlsx')
        E.write_xlsx(r, t, out, ds)
        import openpyxl
        wb = openpyxl.load_workbook(out)
        self.assertEqual(wb.sheetnames, ['raw', 'reduction', 'table'])
        ws = wb['reduction']
        self.assertTrue(str(ws['D2'].value).startswith('=B2/C2'))          # live formulas
        self.assertTrue(str(ws['B3'].value).startswith('=raw!'))          # CaO reads the raw mean; UO3 (converted) is a value


from pxrd_review import epma as EP


class ICDDReplication(unittest.TestCase):
    """The Analysis field of an ICDD entry: mean wt% + the published formula, re-reduced."""
    FLUORMACRAEITE = ('Microprobe analysis, average of 8 (wt.%): K2O 4.41, MnO 12.49, MgO 1.07, Al2O3 0.65, Fe2O3 14.74, TiO2 8.19, '
                      'P2O5 28.51, F 1.55, H2O(calc) 27.13 :[ K0.14 ( H2 O )0.76 ]sigma0.90 [ K0.79 ( H2 O )0.21 ]sigma1.00 '
                      '( Mn1.75 +2 Mg0.25 )sigma2.00 ( Fe1.84 +3 Al0.13 Ti1.02 +4 Mg0.01 )sigma3.00 ( P O4 )4 [ O0.94 F0.81 ( O H )0.25 ]sigma2.00 ( H2 O )10 !3.9 H2 O.')

    def test_parse_analysis_and_formula(self):
        wt, f, n, issues = EP.parse_icdd_analysis(self.FLUORMACRAEITE)
        self.assertEqual((n, wt['K2O'], wt['H2O'], issues), (8, 4.41, 27.13, []))
        counts, ox, fi = EP.parse_icdd_formula(f)
        self.assertAlmostEqual(counts['Mn'], 1.75); self.assertAlmostEqual(counts['P'], 4.0); self.assertAlmostEqual(counts['F'], 0.81)
        self.assertAlmostEqual(counts['H'], 2 * (0.76 + 0.21 + 10 + 3.9) + 0.25, places=6)   # water and hydroxyl H
        self.assertEqual((ox['Mn'], ox['Fe']), (2, 3)); self.assertEqual(fi, [])
        # the notation's traps: ')S2' multipliers, 'Fe3 C1.01' charges, 'OH1.09' groups, vacancies, bare sums, sulfur
        c, ox, fi = EP.parse_icdd_formula('( U0.996 O2 )S2 ( S1.002 O4 )S4 ( H1.998 O )S8', has_sulfur=True)
        self.assertAlmostEqual(c['U'], 1.992); self.assertAlmostEqual(c['S'], 4.008); self.assertAlmostEqual(c['H'], 15.984)
        c, ox, fi = EP.parse_icdd_formula('( Fe3 C1.01 Al0.01 )SI1.02 [O5.91 OH1.09 ]7.00 ( K0.65 ?0.35 )S1.00')
        self.assertEqual((ox['Fe'], round(c['Fe'], 2), round(c['H'], 2), round(c['O'], 2)), (3, 1.01, 1.09, 7.0)); self.assertEqual(fi, [])
        c, ox, fi = EP.parse_icdd_formula('Cu11.75 [ ( As1.44 Sb0.43 )S2.00 Te2.00 ] S13.11', has_sulfur=True)
        self.assertAlmostEqual(c['S'], 13.11)
        c, ox, fi = EP.parse_icdd_formula('( Zn0.32 Fe20.13 +2 )S3.00')
        self.assertTrue(any('group sums' in x for x in fi))                    # a garbled count is caught by the stated sum

    def test_check_reproduces_and_flags(self):
        lines, r = EP.check_analysis(self.FLUORMACRAEITE)
        self.assertIsNotNone(r); self.assertEqual(r['diffs'], []); self.assertLess(r['score'], 0.02)
        # a constituent in the formula but not in the wt% list, a value dropped, a zero for O, a duplicate
        lines, r = EP.check_analysis('Microprobe analysis (wt.%): CaO 7.82, Na2O 0.88, MgO 6.83, Mn2O3 36.55, Fe2O3 1.26, P2O5 22.95, '
                                     'CO2(calc) 7.11, H2O(calc) 14.48: ( Ca0.87 Na0.18 )Sigma1.05 Mg1.05 ( Mn2.87 +3 Fe0.10 +3 )Sigma2.97 O1.93 ( P O4 )2.01 C O3 F1.04 !4.99 H2 O.')
        self.assertEqual([d[0] for d in r['diffs']], ['F'])
        lines, r = EP.check_analysis('Microprobe analysis, average of 18 (wt.%): Nb205 67.34, MgO 21.66, FeO 2.89, MnO 1.18, TiO 0.21: '
                                     '( Mg6.31 Fe0.53 Mn0.20 )S7.03 ( Nb5.96 Ti0.03 )S5.99 O18 ( O H )8.')
        self.assertTrue(any('Nb205' in x for x in r['issues'])); self.assertTrue(any('TiO' in x for x in r['issues']))
        lines, r = EP.check_analysis('Microprobe analysis (wt.%): Na2O 10.31, SiO2 42.19, K2O 0.09, K2O 0.09, ZrO2 41.87, H2O(calc) 3.48: '
                                     '(Na0.94 K0.01)$SI0.95 Zr0.96 Si2.00 [O5.91 OH1.09 ]7.00')
        self.assertIn('K2O listed twice', r['issues'])
        # an O=F line is a deduction, not fluorine; the real F entry still counts
        wt, f, n, issues = EP.parse_icdd_analysis('Microprobe analysis (wt.%): CaO 55.0, P2O5 42.0, F 3.5, O=F 1.47, total 99.0: Ca5.0 P3.0 O12 F0.94.')
        self.assertEqual((wt['F'], 'O' in wt, issues), (3.5, False, []))
        # sulfides: elements in the wt% list, the cation total as the basis
        lines, r = EP.check_analysis('Microprobe analysis, average of 10 (wt.%): Tl 47.41, Cu 15.46, Ag 0.15, As 17.36, Sb 0.41, S 19.20: Tl1.94 Cu2.04 Ag0.01 As1.95 Sb0.03 S5.03.')
        self.assertEqual(r['diffs'], []); self.assertEqual(r['basis'], ('cations', 6.0))
        # an ideal formula cannot be replicated
        lines, r = EP.check_analysis('Microprobe analysis (wt.%): PbO 26.20, ZnO 1.96: Pb2 Zn O2 ( As O4 )4 ( O H )6 !12 H2 O.')
        self.assertIsNone(r); self.assertIn('ideal', lines[0])


if __name__ == '__main__':
    unittest.main()
